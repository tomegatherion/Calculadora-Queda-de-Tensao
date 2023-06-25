import sys
import openpyxl
from PySide6.QtWidgets import (
    QApplication, QLabel, QLineEdit, QPushButton,
    QGridLayout, QTableWidget, QTableWidgetItem,
    QVBoxLayout, QWidget, QGroupBox, QHBoxLayout,
    QProgressBar, QComboBox, QHeaderView, QSpinBox,
    QDoubleSpinBox, QFileDialog
)

css_file = 'style1.css'


# ////////////////////////////////////////////////////////////////////////////////
# Criando a classe principal

class CalculoQuedaTensão(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle('Calculadora \u0394V')

        # ////////////////////////////////////////////////////////////////////////
        # ------------------------------------------------------------------------
        # Grupo para os campos de entrada do usuário
        input_group = QGroupBox('Dados do Circuito')

        # ------------------------------------------------------------------------
        # Labels de identificação dos campos de entrada
        self.nome_label = QLabel('Nome:')
        self.esquema_ligação_label = QLabel('Ligação:')
        self.tensão_entrada_label = QLabel('Tensão de Entrada:')
        self.tensão_m_label = QLabel('Tensão a Montante:')
        self.fator_potência_label = QLabel('Fator de Potência:')
        self.potência_ativa_label = QLabel('Potência Ativa:')
        self.comprimento_condutor_label = QLabel('Comprimento do \nCondutor:')
        self.numero_trifólios_label = QLabel('Numero de Trifólios:')
        self.seção_condutor_label = QLabel('Seção do Condutor:')
        self.material_cabo_label = QLabel('Material do Condutor:')
        self.limite_queda_label = QLabel('Limite \u0394V')

        # ------------------------------------------------------------------------
        # Campos de entrada
        self.nome_edit = QLineEdit()

        self.esquema_ligação_combo = QComboBox()
        self.esquema_ligação_combo.addItems(['1FNT', '2FNT', '3FNT'])

        self.tensão_entrada_edit = QDoubleSpinBox()
        self.tensão_entrada_edit.setMaximum(999999999)
        self.tensão_entrada_edit.setSpecialValueText(" ")

        self.tensão_m_edit = QDoubleSpinBox()
        self.tensão_m_edit.setMaximum(999999999)
        self.tensão_m_edit.setSpecialValueText(" ")

        self.fator_potência_edit = QDoubleSpinBox()
        self.fator_potência_edit.setMinimum(0)
        self.fator_potência_edit.setMaximum(1)
        self.fator_potência_edit.setSpecialValueText(" ")
        self.fator_potência_edit.setDecimals(3)

        self.potência_ativa_edit = QDoubleSpinBox()
        self.potência_ativa_edit.setMaximum(999999999)
        self.potência_ativa_edit.setSpecialValueText(" ")

        self.comprimento_condutor_edit = QDoubleSpinBox()
        self.comprimento_condutor_edit.setMaximum(999999999)
        self.comprimento_condutor_edit.setSpecialValueText(" ")

        self.numero_trifólios_box = QSpinBox()
        self.numero_trifólios_box.setMinimum(1)

        self.seção_condutor_combo = QComboBox()
        self.seção_condutor_combo.addItems(
            [
                '0.5',
                '0.75',
                '1.5',
                '2.5',
                '4',
                '6',
                '10',
                '16',
                '25',
                '35',
                '50',
                '70',
                '95',
                '120',
                '150',
                '185',
                '240',
                '300',
                '400',
                '500',
                '630',
                '800',
                '1000'
            ]
        )

        self.material_cabo_combo = QComboBox()
        self.material_cabo_combo.addItems(['Cobre', 'Alumínio'])

        self.limite_queda_combo = QComboBox()
        self.limite_queda_combo.addItems(['4%', '7%'])

        # ------------------------------------------------------------------------
        # Posicionamento e organização do grupo de entrada
        input_layout = QGridLayout()

        input_layout.addWidget(self.nome_label, 0, 0)
        input_layout.addWidget(self.nome_edit, 0, 1)

        input_layout.addWidget(self.esquema_ligação_label, 1, 0)
        input_layout.addWidget(self.esquema_ligação_combo, 1, 1)

        input_layout.addWidget(self.tensão_entrada_label, 2, 0)
        input_layout.addWidget(self.tensão_entrada_edit, 2, 1)

        input_layout.addWidget(self.tensão_m_label, 3, 0)
        input_layout.addWidget(self.tensão_m_edit, 3, 1)

        input_layout.addWidget(self.fator_potência_label, 4, 0)
        input_layout.addWidget(self.fator_potência_edit, 4, 1)

        input_layout.addWidget(self.potência_ativa_label, 5, 0)
        input_layout.addWidget(self.potência_ativa_edit, 5, 1)

        input_layout.addWidget(self.comprimento_condutor_label, 6, 0)
        input_layout.addWidget(self.comprimento_condutor_edit, 6, 1)

        input_layout.addWidget(self.numero_trifólios_label, 7, 0)
        input_layout.addWidget(self.numero_trifólios_box, 7, 1)

        input_layout.addWidget(self.seção_condutor_label, 8, 0)
        input_layout.addWidget(self.seção_condutor_combo, 8, 1)

        input_layout.addWidget(self.material_cabo_label, 9, 0)
        input_layout.addWidget(self.material_cabo_combo, 9, 1)

        input_layout.addWidget(self.limite_queda_label, 10, 0)
        input_layout.addWidget(self.limite_queda_combo, 10, 1)

        input_group.setLayout(input_layout)

        # ////////////////////////////////////////////////////////////////////////
        # ------------------------------------------------------------------------
        # Grupo para os campos de saída do programa
        output_group = QGroupBox('Resultados')

        # ------------------------------------------------------------------------
        # Labels de identificação dos campos de saída
        self.potência_aparente = QLabel('Potência Aparente:  ')
        self.corrente_label = QLabel('Corrente:  ')
        self.tensão_jusante = QLabel('Tensão a Jusante:  ')
        self.queda_100_label = QLabel('\u0394% Parcial:  ')
        self.queda_somada_label = QLabel('\u0394% Acumulada:  ')
        self.queda_barra = QLabel('Limite:  ')

        # ------------------------------------------------------------------------
        # Campos de saída
        self.potência_aparente_output = QLabel()
        self.corrente_label_output = QLabel()
        self.tensão_jusante_output = QLabel()
        self.queda_100_label_output = QLabel()
        self.queda_somada_output = QLabel()
        self.queda_barra_output = QProgressBar()
        self.queda_barra_output.setFormat('')

        # ------------------------------------------------------------------------
        # Posicionamento e organização do grupo de saída
        output_layout = QGridLayout()

        output_layout.addWidget(self.potência_aparente, 0, 0)
        output_layout.addWidget(self.potência_aparente_output, 0, 1)

        output_layout.addWidget(self.corrente_label, 1, 0)
        output_layout.addWidget(self.corrente_label_output, 1, 1)

        output_layout.addWidget(self.tensão_jusante, 2, 0)
        output_layout.addWidget(self.tensão_jusante_output, 2, 1)

        output_layout.addWidget(self.queda_100_label, 3, 0)
        output_layout.addWidget(self.queda_100_label_output, 3, 1)

        output_layout.addWidget(self.queda_somada_label, 4, 0)
        output_layout.addWidget(self.queda_somada_output, 4, 1)

        output_layout.addWidget(self.queda_barra, 5, 0)
        output_layout.addWidget(self.queda_barra_output, 5, 1)

        output_group.setLayout(output_layout)

        # ------------------------------------------------------------------------
        # Configurações dos botões
        self.calcular_button = QPushButton("Calcular")
        self.calcular_button.clicked.connect(self.calcular_queda_tensão)

        self.salvar_button = QPushButton('Salvar')
        self.salvar_button.clicked.connect(self.salvar_tabela)

        self.histórico_button = QPushButton("Histórico")
        self.histórico_button.clicked.connect(self.toggle_histórico)

        self.exportar_button = QPushButton("Exportar")
        self.exportar_button.clicked.connect(self.exportar_tabela)

        # Botão de acesso ao tema
        self.tema_button = QPushButton('Tema')
        self.tema_button.clicked.connect(self.trocar_estilo)

        # ------------------------------------------------------------------------
        # Tabela de histórico
        self.histórico_table = QTableWidget()
        self.histórico_table.hide()
        self.histórico_table.setColumnCount(14)
        self.histórico_table.setHorizontalHeaderLabels(
            ['Nome',
                'Ligação',
                'V Montante',
                'Fator de\nPotência',
                'Potência\nAtiva',
                'Potência\n Aparente',
                'Comprimento\ndo Condutor',
                'Nº de\nTrifólios',
                'Seção',
                'Material\ndo Cabo',
                'Corrente',
                '\u0394% Parcial',
                '\u0394% Total',
                'Tensão\nFinal']
        )
        header = self.histórico_table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeToContents)

        # ////////////////////////////////////////////////////////////////////////
        # ------------------------------------------------------------------------
        # Layout principal
        layout = QVBoxLayout()
        layout.addWidget(input_group)
        layout.addWidget(output_group)
        layout.addWidget(self.calcular_button)
        layout.addWidget(self.salvar_button)

        # ------------------------------------------------------------------------
        # Layout secundário para os botões
        button_layout = QHBoxLayout()
        button_layout.addWidget(self.tema_button)
        button_layout.addWidget(self.histórico_button)
        button_layout.addWidget(self.exportar_button)
        layout.addLayout(button_layout)

        # ------------------------------------------------------------------------
        # Configura a quantidade de linhas visíveis na tabela
        num_linhas_visíveis = 8  # Altere o número de linhas visíveis aqui
        altura_linha = self.histórico_table.verticalHeader().\
            defaultSectionSize()

        altura_tabela = num_linhas_visíveis * altura_linha
        self.histórico_table.setFixedHeight(altura_tabela)

        # ------------------------------------------------------------------------
        # Adicionar tabela de histórico
        layout.addWidget(self.histórico_table)

        # ------------------------------------------------------------------------
        # Definir altura fixa para o layout principal
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(10)
        self.setLayout(layout)
        self.setFixedHeight(self.minimumSizeHint().height())

    # ///////////////////////////////////////////////////////////////////////////////////////////////////////////////////
    # ------------------------------------------------------------------------------------------------------------------
    # Funções

    # Função de calculo principal
    def calcular_queda_tensão(self):
        ligação = self.esquema_ligação_combo.currentText()
        tensão_entrada = self.tensão_entrada_edit.value()
        tensão_montante = self.tensão_m_edit.value()
        fp = self.fator_potência_edit.value()
        pot_w = self.potência_ativa_edit.value()
        pot_va = pot_w / fp
        if ligação == "1FNT" or ligação == "2FNT":
            fator_k = 200
            impedância = (tensão_entrada ** 2) / (pot_w / fp)
            corrente = tensão_montante / impedância

        else:
            fator_k = 173.2
            impedância = ((tensão_entrada * 1.73205080757) ** 2) / (pot_w / fp)
            corrente = (tensão_montante * 1.73205080757) / impedância

        dist = self.comprimento_condutor_edit.value()
        n_trifólios = self.numero_trifólios_box.value()
        bitola = float(self.seção_condutor_combo.currentText())
        cabo = self.material_cabo_combo.currentText()
        if self.limite_queda_combo.currentText() == '4%':
            limite_qt = 4
        else:
            limite_qt = 7

        if cabo == 'Cobre':
            resistividade = 0.0178571428571429
        else:
            resistividade = 0.0292056074766355

        queda_100 = ((fator_k * resistividade * dist * corrente) /
                     ((n_trifólios * bitola) * tensão_montante))

        queda_v = (tensão_montante * (queda_100 / 100))
        tensão_jusante = tensão_montante - queda_v

        queda_somada = ((tensão_entrada - tensão_jusante)
                        * 100) / tensão_entrada

        self.potência_aparente_output.setText("{:.2f} VA".format(pot_va))
        self.corrente_label_output.setText("{:.2f} A".format(corrente))
        self.tensão_jusante_output.setText("{:.2f} V".format(tensão_jusante))
        self.queda_100_label_output.setText("{:.2f} %".format(queda_100))
        self.queda_somada_output.setText("{:.2f} %".format(queda_somada))

        self.queda_barra_output.setRange(0, (limite_qt))
        self.queda_barra_output.setValue(queda_100)
        if queda_100 > limite_qt:
            self.queda_barra_output.setStyleSheet(
                "QProgressBar::chunk { background-color: red; }")
        else:
            self.queda_barra_output.setStyleSheet(
                "QProgressBar::chunk { background-color: #337ab7; }")

    # ------------------------------------------------------------------------
    # Função para trocar o arquivo CSS carregado
    def trocar_estilo(self):
        global css_file
        if css_file == 'style1.css':
            css_file = 'style2.css'
        elif css_file == 'style2.css':
            css_file = 'style3.css'
        elif css_file == 'style3.css':
            css_file = 'style1.css'

        # Carregar o arquivo CSS atualizado
        with open(f'styles/{css_file}', 'r') as file:
            estilo = file.read()
        app.setStyleSheet(estilo)

    # ------------------------------------------------------------------------------------------------------------------
    # Função para salvar os cálculos na tabela de histórico

    # declaração de variáveis
    def salvar_tabela(self):
        row_count = self.histórico_table.rowCount()
        nome = self.nome_edit.text()
        ligação = self.esquema_ligação_combo.currentText()
        tensão_entrada = str(self.tensão_entrada_edit.value()) + ' V'
        fp = str(self.fator_potência_edit.value())
        pot_w = str(self.potência_ativa_edit.value()) + ' W'
        potência_aparente = self.potência_aparente_output.text()
        dist = str(self.comprimento_condutor_edit.value()) + ' m'
        n_trifólios = str(self.numero_trifólios_box.value()) + ' x'
        bitola = str(self.seção_condutor_combo.currentText()) + ' mm²'
        cabo = self.material_cabo_combo.currentText()
        corrente = self.corrente_label_output.text()
        queda_100 = self.queda_100_label_output.text()
        queda_somada = self.queda_somada_output.text()
        tensão_jusante = self.tensão_jusante_output.text()

        # Inserção de colunas
        self.histórico_table.insertRow(row_count)
        self.histórico_table.setItem(row_count, 0, QTableWidgetItem(nome))
        self.histórico_table.setItem(row_count, 1, QTableWidgetItem(ligação))
        self.histórico_table.setItem(row_count, 2, QTableWidgetItem
                                     (tensão_entrada))

        self.histórico_table.setItem(row_count, 3, QTableWidgetItem(fp))
        self.histórico_table.setItem(row_count, 4, QTableWidgetItem(pot_w))
        self.histórico_table.setItem(row_count, 5, QTableWidgetItem
                                     (potência_aparente))

        self.histórico_table.setItem(row_count, 6, QTableWidgetItem(dist))
        self.histórico_table.setItem(row_count, 7, QTableWidgetItem
                                     (n_trifólios))

        self.histórico_table.setItem(row_count, 8, QTableWidgetItem(bitola))
        self.histórico_table.setItem(row_count, 9, QTableWidgetItem(cabo))
        self.histórico_table.setItem(row_count, 10, QTableWidgetItem(corrente))
        self.histórico_table.setItem(row_count, 11, QTableWidgetItem
                                     (queda_100))

        self.histórico_table.setItem(row_count, 12, QTableWidgetItem
                                     (queda_somada))

        self.histórico_table.setItem(row_count, 13, QTableWidgetItem
                                     (tensão_jusante))

    # ------------------------------------------------------------------------------------------------------------------
    # Muda a visibilidade da área da tabela de histórico
    def toggle_histórico(self):
        if self.histórico_table.isVisible():
            self.histórico_table.hide()
            self.setFixedHeight(self.minimumSizeHint().height())
        else:
            self.histórico_table.show()
            self.setFixedHeight(self.minimumSizeHint().height())

    # ------------------------------------------------------------------------------------------------------------------
    # Exporta a tabela para um arquivo
    def exportar_tabela(self):
        # Criar um novo arquivo Excel
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Obter os cabeçalhos da tabela de histórico
        headers = []
        for col in range(self.histórico_table.columnCount()):
            header = self.histórico_table.horizontalHeaderItem(col)
            if header is not None:
                headers.append(header.text())
            else:
                headers.append("")

        # Escrever os cabeçalhos na primeira linha da planilha
        for col, header in enumerate(headers):
            sheet.cell(row=1, column=col + 1, value=header)

        # Obter os dados da tabela de histórico
        for row in range(self.histórico_table.rowCount()):
            for col in range(self.histórico_table.columnCount()):
                item = self.histórico_table.item(row, col)
                if item is not None:
                    value = item.text()
                    sheet.cell(row=row + 2, column=col + 1, value=value)

        # Abrir uma janela de diálogo para salvar o arquivo
        file_dialog = QFileDialog()
        file_dialog.setDefaultSuffix("xlsx")
        options = QFileDialog.Options()
        file_path, _ = file_dialog.getSaveFileName(
            None, "Salvar Arquivo", "", "Arquivos Excel (*.xlsx)",
            options=options
        )

        if file_path:
            # Salvar o arquivo Excel
            workbook.save(file_path)


if __name__ == '__main__':
    app = QApplication(sys.argv)

    # Carregar o arquivo CSS externo
    # with open('styles/style1.css', 'r') as file:
    #    estilo = file.read()
    with open(f'styles/{css_file}', 'r') as file:
        estilo = file.read()
    app.setStyleSheet(estilo)

    # app.setStyleSheet(estilo)

    window = CalculoQuedaTensão()
    window.show()
    sys.exit(app.exec())
